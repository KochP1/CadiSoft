import pandas as pd
import io
from flask import send_file
from datetime import datetime
from openpyxl.styles import PatternFill, Font

def planilla_calificaciones(planilla, curso, seccion):
        # Crear DataFrame
        df = pd.DataFrame(planilla)
        
        # Reordenar columnas para mejor presentación
        column_order = ['nombre', 'segundoNombre', 'apellido', 'segundoApellido',  'cedula', 'logro_1', 'logro_2', 'logro_3', 'logro_4', 'logro_5', 'definitiva', 'fecha_inscripcion', 'fecha_expiracion']
        df = df[column_order]
        
        # Renombrar columnas en español
        df.columns = ['Nombre', 'Segundo nombre', 'Apellido', 'Segundo apellido', 'Cédula', 'Logro 1', 'Logro 2', 'Logro 3', 'Logro 4', 'Logro 5', 'Definitiva', 'Fecha Inscripción', 'Fecha Expiración']
        
        # Crear archivo Excel en memoria
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Planilla', index=False)
            
            # Obtener el libro y hoja de trabajo para formatear
            workbook = writer.book
            worksheet = writer.sheets['Planilla']
            
            # Ajustar ancho de columnas automáticamente
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Formato de header CORREGIDO
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            # Aplicar formato al header (fila 1)
            for row in worksheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.fill = header_fill
                    cell.font = header_font
        
        output.seek(0)
        
        # Crear nombre de archivo con timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"planilla_inscripciones_{curso}_{seccion}.xlsx"
        
        # Retornar archivo para descarga
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

def carga_masiva_calificaciones(archivo):
    if archivo.filename.lower().endswith('.csv'):
        df = pd.read_csv(archivo, header=None)
    else:
        df = pd.read_excel(archivo, header=None)
    
    if df.empty:
        return []
    
    headers = df.iloc[0].tolist()
    notas = [dict(zip(headers, row)) for row in df.iloc[1:].values]
    
    return notas