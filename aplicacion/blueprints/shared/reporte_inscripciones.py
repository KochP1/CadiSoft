import pandas as pd
import io
from flask import send_file
from aplicacion.blueprints.cursos.routes import dateToString
from datetime import datetime
from openpyxl.styles import PatternFill, Font

def reporte_inscripciones(db, tipo, p_cedula, p_fecha_inscripcion, p_idSeccion, p_fecha_expiracion, p_es_activa):
    with db.cursor() as cur:
        cur.callproc('reporte_inscripciones_sp', [p_cedula, p_fecha_inscripcion, p_idSeccion, p_fecha_expiracion, tipo, p_es_activa])
        result = cur.fetchall()
        columNames = [column[0] for column in cur.description]
        
        inscripciones_dict = {}
        
        for record in result:
            record_dict = dict(zip(columNames, record))
            
            key = f"{record_dict['idusuarios']}_{record_dict['fecha_inscripcion']}_{record_dict['fecha_expiracion']}"
            
            if key not in inscripciones_dict:
                inscripciones_dict[key] = {
                    'nombre': record_dict['nombre'],
                    'apellido': record_dict['apellido'],
                    'cedula': record_dict['cedula'],
                    'fecha_inscripcion': record_dict['fecha_inscripcion'],
                    'fecha_expiracion': record_dict['fecha_expiracion'],
                    'tipo': record_dict['tipo'],
                    'status': record_dict['status']
                }
        
        inscripciones = list(inscripciones_dict.values())

        for record in inscripciones:
            record['fecha_inscripcion'] = dateToString(record['fecha_inscripcion'])
            record['fecha_expiracion'] = dateToString(record['fecha_expiracion'])
        
        # Crear DataFrame
        df = pd.DataFrame(inscripciones)
        
        # Reordenar columnas para mejor presentación
        column_order = ['nombre', 'apellido', 'cedula', 'fecha_inscripcion', 'fecha_expiracion', 'tipo', 'status']
        df = df[column_order]
        
        # Renombrar columnas en español
        df.columns = ['Nombre', 'Apellido', 'Cédula', 'Fecha Inscripción', 'Fecha Expiración', 'Tipo', 'Estado']
        
        # Convertir estado a texto más legible
        df['Estado'] = df['Estado'].apply(lambda x: 'Activa' if x else 'Inactiva')
        
        # Crear archivo Excel en memoria
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Inscripciones', index=False)
            
            # Obtener el libro y hoja de trabajo para formatear
            workbook = writer.book
            worksheet = writer.sheets['Inscripciones']
            
            # Ajustar ancho de columnas automáticamente
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Formato de header CORREGIDO
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            # Aplicar formato al header (fila 1)
            for row in worksheet.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.fill = header_fill
                    cell.font = header_font
        
        output.seek(0)
        
        # Crear nombre de archivo con timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"reporte_inscripciones_{timestamp}.xlsx"
        
        # Retornar archivo para descarga
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )